import calendar
import json
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import holidays

DEFAULT_FILE = Path(__file__).parent / "B_CM_Audit.xlsx"
TRACKER_CONFIG_FILE = Path(__file__).parent / "tracker_config.json"
KR_HOLIDAYS = holidays.KR()  # 대한민국 공휴일(대체공휴일 포함), 연도는 조회 시점에 맞춰 자동 확장됨

# 열 인덱스 (1-based)
COL_TRACKER_NAME  = 1   # A
COL_TRACKER_TYPE  = 2   # B
COL_ITEM_COUNT    = 4   # D
COL_FILE_NAME     = 5   # E
COL_FIRST_EDIT    = 6   # F
COL_STATUS        = 8   # H
COL_CURRENT_VERSION = 9  # I ("현재 버전" - 문서 자체의 버전)
COL_PR_ID         = 10  # J
COL_VERSIONING    = 11  # K
COL_VER_DESC      = 12  # L
COL_CREATE_DATE_CURRENT = 19  # S ("Create Date 최신 여부" - B_Audit_Data_Creation.py에서 계산)
COL_TARGET_VERSION = 20  # T ("리뷰레포트 기재 버전" - Review Report에 기재된 대상 문서 버전)
COL_VERSION_CHECK_FAIL_REASON = 21  # U (버전 자동 인식 실패 사유)
COL_SAVE_RULE     = 22  # V
COL_VERSION_RULE  = 23  # W
COL_DOC_HISTORY_RULE = 24  # X
COL_STATUS_RULE   = 25  # Y
COL_COMMENT       = 26  # Z

PROCESS_TAG_RE = re.compile(r"^\[[A-Z]+\.\d+[A-Z]?\]") # [SAF.2]Functional Safety Audit Report -> [SAF.2]
EMPTY_FILE_VALUES = {"", "0", "미업로드"}
EMPTY_DATE_VALUES = {"", "미업로드"}
VERSION_DEADLINE_DAYS = 10 # 2.1) 버전 규칙 준수 : 첫 Edit 이후, Working Day 10일 이내 Version Up이 진행되지 않은 상태(26.05.28)


def _load_tracker_config():
    """주기적/이벤트성 산출물 트래커 목록을 tracker_config.json에서 읽어온다.

    새 산출물이 프로젝트에 추가됐을 때 이 코드를 건드리지 않고 그 파일만 수정하면 되도록
    목록(데이터)을 로직(코드)에서 분리했다.
    """
    if not TRACKER_CONFIG_FILE.exists():
        raise FileNotFoundError(
            f"{TRACKER_CONFIG_FILE} 파일을 찾을 수 없습니다. "
            "주기적/이벤트성 산출물 트래커 목록이 이 파일에 있어야 합니다."
        )
    with TRACKER_CONFIG_FILE.open(encoding="utf-8") as f:
        data = json.load(f)
    return set(data.get("periodic_trackers", [])), set(data.get("eventbased_trackers", []))


# 이벤트성 활동 산출물 목록 / 주기적 활동 산출물 목록 (tracker_config.json에서 로드)
# 2.2)/2.3) 버전규칙준수 : 주기적으로 업로드 되는 산출물 또는 일정한 주기 없이 업데이트되는 산출물의 경우,
# 지정된 주기 내 혹은 Released/Read Only 상태임에도 불구하고 Create Date가 진행되지 않은 상태(26.05.28)
PERIODIC_TRACKERS, EVENTBASED_TRACKERS = _load_tracker_config()

# Create Date 검사 대상 상태 (이 워크플로우를 타는 산출물은 Released 또는 Read Only가 최종 상태이고,
# Create Date를 누르면 즉시 이 상태로 자동 복귀하기 때문에 "현재 상태"만으로는 최신 여부를 알 수 없다.
# 그래서 이 상태일 때만 COL_CREATE_DATE_CURRENT(B_Audit_Data_Creation.py에서 계산한 값)로 최신 여부를 판정한다.
EVENTBASED_TERMINAL_STATUSES = {"Released", "Read Only"}

# 주기적 활동 산출물의 마감 주기 - 프로젝트마다 다르므로 run() 호출 시 지정한다 (기본값: 2주 단위, 월요일)
WEEKDAY_KR = {"월": 0, "화": 1, "수": 2, "목": 3, "금": 4, "토": 5, "일": 6}
DEFAULT_PERIODIC_CADENCE = "biweekly"  # "weekly" | "biweekly" | "monthly"
DEFAULT_PERIODIC_ANCHOR = 0  # weekly/biweekly: 요일(0=월~6=일), monthly: 일자(1~31)


def _clamp_day(year: int, month: int, day: int) -> date:
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, min(day, last_day))


def compute_next_periodic_due(last_create_date: datetime, cadence: str, anchor: int) -> date:
    """마지막 Create Date 이후, 선택한 주기(요일/일자)가 처음으로 돌아오는 마감일을 계산한다."""
    if cadence == "weekly":
        d = last_create_date.date() + timedelta(days=1)
        while d.weekday() != anchor:
            d += timedelta(days=1)
        return d

    if cadence == "biweekly":
        d = last_create_date.date() + timedelta(days=14)
        while d.weekday() != anchor:
            d += timedelta(days=1)
        return d

    if cadence == "monthly":
        year, month = last_create_date.year, last_create_date.month
        candidate = _clamp_day(year, month, anchor)
        if candidate <= last_create_date.date():
            month += 1
            if month > 12:
                month = 1
                year += 1
            candidate = _clamp_day(year, month, anchor)
        return candidate

    raise ValueError(f"알 수 없는 주기 종류: {cadence}")


def get_cell_value(ws, row, col):
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def strip_process_tag(tracker_name: str) -> str:
    return PROCESS_TAG_RE.sub("", tracker_name).strip()


def normalize_for_naming_check(text: str) -> str:
    """파일명 규칙 비교용 정규화: 대소문자, 공백/언더스코어 구분자 차이는 같은 것으로 취급한다."""
    return re.sub(r"[_\s]+", " ", text).strip().lower()


def has_file(file_name: str) -> bool:
    return file_name.lower() not in EMPTY_FILE_VALUES


def parse_datetime(value: str):
    if value.lower() in EMPTY_DATE_VALUES:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def business_days_between(start: datetime, end: datetime) -> int:
    """start와 end 사이의 영업일 수(주말·대한민국 공휴일 제외)를 센다."""
    if end <= start:
        return 0
    total = 0
    current = start.date() + timedelta(days=1)
    end_date = end.date()
    while current <= end_date:
        if current.weekday() < 5 and current not in KR_HOLIDAYS:
            total += 1
        current += timedelta(days=1)
    return total


def append_comment(existing: str, new_reasons: list) -> str:
    new_text = " / ".join(new_reasons)
    if not existing:
        return new_text
    if new_text in existing:
        return existing
    return existing + " / " + new_text


# ── 저장 규칙 검사 ─────────────────────────────────────────────────────────────

def check_save_rule(ws, row):
    tracker_type = get_cell_value(ws, row, COL_TRACKER_TYPE)
    tracker_name = get_cell_value(ws, row, COL_TRACKER_NAME)
    item_count_raw = get_cell_value(ws, row, COL_ITEM_COUNT)
    file_name = get_cell_value(ws, row, COL_FILE_NAME)

    try:
        item_count = int(float(item_count_raw)) if item_count_raw else 0
    except ValueError:
        item_count = 0

    file_exists = has_file(file_name)

    if item_count == 0 and not file_exists:
        return None

    if not file_exists and tracker_type != "Document":
        return None

    reasons = []
    detail_reasons = []  # codebeamer엔 안 보내고 엑셀 요약 시트에만 보여줄 상세 사유

    if tracker_type == "Document" and item_count >= 2: # 1.2) 저장 규칙 준수 : Document Tracker의 경우, 하나의 Tracker 내 하나의 파일만 등재되었는지 확인(26.05.28)
        reason = f"파일 {item_count}개 등재됨"
        reasons.append(reason)
        detail_reasons.append(reason)

    if file_exists:
        pure_name = strip_process_tag(tracker_name)
        file_stem = Path(file_name).stem
        # 대소문자, 공백/언더스코어 표기 차이만 허용. 버전/날짜 등 다른 접미사가 붙으면 안 되므로 완전 일치로 검사한다
        if normalize_for_naming_check(pure_name) != normalize_for_naming_check(file_stem): # 1.3) 저장 규칙 준수 : Configuration Management Plan 내 '4.2.4 File Naming Rule'에 맞게 산출물들이 저장되고 관리되고 있는지 확인(26.05.28)
            reasons.append("File Naming Rule 불일치")
            detail_reasons.append(f"File Naming Rule 불일치 (트래커명: '{pure_name}', 실제 파일명: '{file_stem}')")

    return (len(reasons) == 0, reasons, detail_reasons)


# ── 버전 규칙 검사 ─────────────────────────────────────────────────────────────

def check_version_rule(ws, row):
    first_edit_raw = get_cell_value(ws, row, COL_FIRST_EDIT) # 첫 Edit 시각
    versioning_raw = get_cell_value(ws, row, COL_VERSIONING) # 버저닝 시각

    first_edit = parse_datetime(first_edit_raw)

    if first_edit is None:
        return None

    reasons = []

    versioning = parse_datetime(versioning_raw)

    if versioning is None:
        reasons.append("첫 Edit 이후 버저닝 미수행")
    else:
        business_days = business_days_between(first_edit, versioning)
        if business_days > VERSION_DEADLINE_DAYS: # 2.1) 버전 규칙 준수 : 첫 Edit 이후, Working Day 10일 이내 Version Up이 진행되지 않은 상태(26.05.28)
            reasons.append(
                f"버저닝 지연 (첫 Edit {first_edit.date()} → 버저닝 {versioning.date()}, "
                f"영업일 {business_days}일 경과, 기준 {VERSION_DEADLINE_DAYS}일 초과)"
            )

    return (len(reasons) == 0, reasons, reasons)  # 이미 구체적인 사유라 상세 사유도 동일하게 사용


# ── 주기적 활동 산출물 Create Date 검사 ──────────────────────────────────────

def check_periodic_create_date(ws, row, cadence=DEFAULT_PERIODIC_CADENCE, anchor=DEFAULT_PERIODIC_ANCHOR):
    tracker_name_raw = get_cell_value(ws, row, COL_TRACKER_NAME)
    pure_name = strip_process_tag(tracker_name_raw)

    # 주기적 활동 산출물이 아니면 스킵
    if pure_name not in PERIODIC_TRACKERS:
        return None

    first_edit_raw = get_cell_value(ws, row, COL_FIRST_EDIT)
    # 활동 자체가 없으면(아직 시작 안 한 산출물) 스킵
    if parse_datetime(first_edit_raw) is None:
        return None

    versioning_raw = get_cell_value(ws, row, COL_VERSIONING)
    versioning = parse_datetime(versioning_raw)

    if versioning is None: # 2.2) 버전규칙준수 : 주기적으로 업로드 되는 산출물(날짜 관리)의 경우, 지정된 주기 내 Create Date가 진행되지 않은 상태(26.05.28)
        reasons = ["활동은 있으나 Create Date 미수행"]
        return (False, reasons, reasons)

    due_date = compute_next_periodic_due(versioning, cadence, anchor)
    today = datetime.now().date()

    if today >= due_date:
        reasons = [f"주기적 활동 Create Date 지연 (마지막 Create Date {versioning.date()}, 다음 마감 {due_date} 초과)"]
        return (False, reasons, reasons)

    return (True, [], [])


# ── 이벤트성 산출물 Create Date 검사 ──────────────────────────────────────────

def check_eventbased_create_date(ws, row):
    tracker_name_raw = get_cell_value(ws, row, COL_TRACKER_NAME)
    pure_name = strip_process_tag(tracker_name_raw)

    # 이벤트성 산출물이 아니면 스킵
    if pure_name not in EVENTBASED_TRACKERS:
        return None

    status = get_cell_value(ws, row, COL_STATUS)

    # Released/Read Only가 아니면(아직 진행 중이거나 다른 워크플로우) 스킵
    if status not in EVENTBASED_TERMINAL_STATUSES:
        return None

    # Create Date를 누르면 즉시 Released/Read Only로 자동 복귀("back")하므로, 현재 상태만으로는
    # Create Date를 했는지 알 수 없다. 가장 최근 히스토리 항목이 "back" 전이였는지(=마지막으로 한 일이
    # Create Date였는지, 그 이후 다른 수정이 없었는지)를 B_Audit_Data_Creation.py가 미리 계산해둔 값으로 판정한다.
    create_date_current = get_cell_value(ws, row, COL_CREATE_DATE_CURRENT).strip().upper() == "TRUE"

    if not create_date_current: # 2.3) 버전규칙준수 : 일정한 주기 없이 업데이트되는 산출물의 경우, Released/Read Only 상태임에도 불구하고 마지막 수정 이후 Create Date가 진행되지 않은 상태(26.05.28)
        reasons = [f"상태 '{status}'이나 마지막 수정 이후 Create Date 미수행"]
        return (False, reasons, reasons)

    return (True, [], [])


# ── 문서 이력 기술 규칙 검사 ──────────────────────────────────────────────────
# 3.1) 문서이력기술규칙 준수 : Configuration Management Plan 내 '4.2.6 History Description Rule'에 기재한 이력 작성 규칙에 따라 작성되었는지 확인(단, Approved 상태의 버전은 문서 이력을 확인하지 않는다)
PR_ID_SPLIT_RE = re.compile(r"[\s,]+")

PR_IN_DESC_RE = re.compile(r"\bPR[^\d]*?(\d+)", re.IGNORECASE) # 하이픈 유무와 무관하게 PR 뒤에 오는 첫 숫자를 PR 번호로 인식 (PR-04, PR~-04, PR04, PR 04 모두 허용)

def check_doc_history_rule(ws, row):
    status = get_cell_value(ws, row, COL_STATUS)

    if status == 'Approved': # Approved 상태의 버전은 문서 이력 기술 규칙 전체를 검사하지 않는다(26.06.08)
        return None

    pr_id_raw = get_cell_value(ws, row, COL_PR_ID)

    pr_ids = set()
    for token in PR_ID_SPLIT_RE.split(pr_id_raw):
        token = token.strip()
        if token.isdigit():
            pr_ids.add(int(token))

    reasons = []
    ver_desc_raw = get_cell_value(ws, row, COL_VER_DESC)
    if not ver_desc_raw:
        reasons.append("버전 이력 Description이 작성되지 않음") # 3.1) 문서이력기술규칙 준수 : Configuration Management Plan 내 '4.2.6 Hisory Description Rule'에 기재한 이력 작성 규칙에 따라 작성되었는지 확인

    desc_pr_ids = set()
    for m in PR_IN_DESC_RE.finditer(ver_desc_raw):
        desc_pr_ids.add(int(m.group(1)))

    missing = pr_ids - desc_pr_ids 
    

    if missing: # 3.2) 문서이력기술규칙 준수 : PR/CR로 인해 수정된 경우, 해당 산출물의 수정사항에 맞는 PR/CR-ID가 기입되었는지 확인(26.05.28)
        missing_str = ", ".join(str(n) for n in sorted(missing))
        reasons.append(f"PR 조치 미기술: {missing_str}")

    return (len(reasons) == 0, reasons, reasons)


# ── 상태 규칙 검사 ─────────────────────────────────────────────────────────────

UPLOAD_TRUE_STATUSES  = {"Approved", "Internal Baselined", "Gate Baselined", "Waiting for Approval"}
UPLOAD_FALSE_STATUSES = {"In Review", "Open"}

def check_status_rule(ws, row):
    tracker_name_raw = get_cell_value(ws, row, COL_TRACKER_NAME)
    if strip_process_tag(tracker_name_raw) in EVENTBASED_TRACKERS:
        # 이벤트성 산출물은 Released/Read Only + Create Date 워크플로우를 쓰므로
        # 이 규칙(Waiting for Approval 전후 상태 검사) 대상이 아니다
        return None

    upload_raw = get_cell_value(ws, row, 15)  # O열: 리뷰레포트 업로드 여부
    status_raw = get_cell_value(ws, row, 8)   # H열: 현재 상태

    upload_val = upload_raw.strip().upper()

    if upload_raw.strip() == "해당없음":
        return (True, [], [])

    if upload_val not in ("TRUE", "FALSE"):
        return None

    reasons = []

    if upload_val == "TRUE":
        if status_raw not in UPLOAD_TRUE_STATUSES:
            reasons.append(f"리뷰레포트 업로드 이후이나 상태가 '{status_raw}'") # 4.2) 상태규칙준수 : 산출물의 Review Report가 업로드 되었을 때, 대상 산출물의 상태가 Waiting for Approval 이후 상태인지 확인(26.05.28)
    else:  # FALSE
        if status_raw not in UPLOAD_FALSE_STATUSES:
            reasons.append(f"리뷰레포트 업로드 이전이나 상태가 '{status_raw}'") # 4.1) 상태규칙준수 : 산출물의 Review Report가 업로드 되기 전, 대상 산출물의 상태가 Waiting for Approval 이전 상태인지 확인(26.05.28)

    return (len(reasons) == 0, reasons, reasons)


# ── 리뷰레포트 대상 버전 규칙 검사 ────────────────────────────────────────────

def check_review_report_version_rule(ws, row):
    """산출물이 Approved 상태일 때, Review Report에 기재된 대상 문서 버전이
    실제 현재 버전과 일치하는지 확인한다. (최신 버전에 대한 리뷰 없이 승인되는 것을 방지)"""
    status_raw = get_cell_value(ws, row, COL_STATUS)
    if status_raw != 'Approved':
        return None

    current_version = get_cell_value(ws, row, COL_CURRENT_VERSION)
    target_version = get_cell_value(ws, row, COL_TARGET_VERSION)
    if not current_version or current_version == '미업로드' or not target_version:
        # 현재 버전을 모르거나, 리뷰레포트가 없거나 대상 버전을 자동 인식하지 못한 경우
        # (후자는 run()에서 별도 "판정 불가" 목록으로 사람에게 안내한다)
        return None

    if target_version.strip() != current_version.strip():
        reasons = [f"Review Report에 기재된 버전({target_version})이 현재 버전({current_version})과 다름"]
        return (False, reasons, reasons)

    return (True, [], [])


# ── 사람이 보기 위한 요약 시트 ──────────────────────────────────────────────────

RULE_STATUS_TEXT = {1: "OK", 2: "NG"}
SUMMARY_RULE_COLUMNS = [
    ("저장 규칙", COL_SAVE_RULE),
    ("버전 규칙", COL_VERSION_RULE),
    ("문서 이력 기술 규칙", COL_DOC_HISTORY_RULE),
    ("상태 규칙", COL_STATUS_RULE),
]


def _build_summary_sheet(wb, detail_ws, row_details=None):
    """detail_ws(재료 데이터 + 규칙 결과 원본)를 바탕으로, 트래커명/규칙 결과(OK,NG)/사유만 보이는
    사람이 읽기 쉬운 요약 시트를 새로 만든다. detail_ws는 D_Result_Update.py가 그대로 읽으므로
    (숫자 1/2가 codebeamer의 실제 선택값 ID) 값 자체는 건드리지 않고 새 시트에만 반영한다.

    row_details가 있으면(codebeamer로는 안 나가는 상세 사유) 그걸 우선 쓰고, 없으면 Comment 열의
    간단한 사유를 그대로 보여준다."""
    row_details = row_details or {}
    if "요약" in wb.sheetnames:
        del wb["요약"]
    summary_ws = wb.create_sheet("요약")  # 맨 뒤에 추가 - pandas가 기본으로 읽는 0번 시트(상세) 순서를 건드리지 않기 위함

    headers = ["트래커명"] + [label for label, _ in SUMMARY_RULE_COLUMNS] + ["Comment"]
    summary_ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D6CDF")
    for cell in summary_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ok_fill = PatternFill("solid", fgColor="E3F5E6")
    ok_font = Font(color="1E7E34", bold=True)
    ng_fill = PatternFill("solid", fgColor="FBE4E2")
    ng_font = Font(color="C0392B", bold=True)

    for row in range(2, detail_ws.max_row + 1):
        tracker_name = get_cell_value(detail_ws, row, COL_TRACKER_NAME)
        if not tracker_name:
            continue

        if row in row_details:
            comment_pretty = "\n".join(f"- {r.strip()}" for r in row_details[row] if r.strip())
        else:
            comment_raw = get_cell_value(detail_ws, row, COL_COMMENT)
            comment_pretty = "\n".join(f"- {r.strip()}" for r in comment_raw.split(" / ") if r.strip())

        summary_ws.append(
            [tracker_name]
            + [RULE_STATUS_TEXT.get(detail_ws.cell(row=row, column=col).value, "-") for _, col in SUMMARY_RULE_COLUMNS]
            + [comment_pretty]
        )

        out_row = summary_ws.max_row
        for i in range(2, 2 + len(SUMMARY_RULE_COLUMNS)):
            cell = summary_ws.cell(row=out_row, column=i)
            cell.alignment = Alignment(horizontal="center")
            if cell.value == "OK":
                cell.fill, cell.font = ok_fill, ok_font
            elif cell.value == "NG":
                cell.fill, cell.font = ng_fill, ng_font

        comment_cell = summary_ws.cell(row=out_row, column=len(headers))
        comment_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if comment_pretty:
            summary_ws.row_dimensions[out_row].height = 14 * (comment_pretty.count("\n") + 1) + 8

    for i, width in enumerate([40, 12, 12, 20, 12, 60], start=1):
        summary_ws.column_dimensions[get_column_letter(i)].width = width

    summary_ws.freeze_panes = "A2"
    summary_ws.auto_filter.ref = summary_ws.dimensions

    wb.active = wb.sheetnames.index("요약")  # 파일을 열었을 때 요약 시트가 먼저 보이도록


def _build_version_check_failure_sheet(wb, failures):
    """Approved 상태인데 Review Report에서 대상 버전을 자동으로 읽지 못한 항목을
    별도 시트로 안내한다 - 잘못 판정하는 대신 사람이 직접 확인하도록."""
    if "판정불가" in wb.sheetnames:
        del wb["판정불가"]
    ws = wb.create_sheet("판정불가")
    ws.append(["트래커명", "사유"])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="E67E22")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for item in failures:
        ws.append([item["트래커명"], item["사유"]])

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40
    ws.freeze_panes = "A2"


# ── 메인 실행 ──────────────────────────────────────────────────────────────────

def run(file_path: str, periodic_cadence: str = DEFAULT_PERIODIC_CADENCE, periodic_anchor: int = DEFAULT_PERIODIC_ANCHOR):
    path = Path(file_path)
    if not path.exists():
        print(f"파일을 찾을 수 없습니다: {file_path}")
        sys.exit(1)

    print(f"\n파일 로딩 중: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = "상세"

    save_ok = save_ng = save_skip = 0
    ver_ok = ver_ng = ver_skip = 0
    periodic_ok = periodic_ng = periodic_skip = 0
    event_ok = event_ng = event_skip = 0
    doc_hist_ok = doc_hist_ng = doc_hist_skip = 0
    status_ok = status_ng = status_skip = 0
    review_ver_ok = review_ver_ng = review_ver_skip = 0
    version_check_failures = []  # 자동으로 리뷰 대상 버전을 인식하지 못한 항목 (사람이 직접 확인해야 함)
    row_details = {}  # row -> 상세 사유 목록 (codebeamer엔 안 보내고 요약 시트에만 노출)

    ws.cell(row=1, column=COL_SAVE_RULE).value        = "저장 규칙"
    ws.cell(row=1, column=COL_VERSION_RULE).value      = "버전 규칙"
    ws.cell(row=1, column=COL_DOC_HISTORY_RULE).value  = "문서 이력 기술 규칙"
    ws.cell(row=1, column=COL_STATUS_RULE).value       = "상태 규칙"
    ws.cell(row=1, column=COL_COMMENT).value           = "Comment"

    for row in range(2, ws.max_row + 1):
        existing_comment = get_cell_value(ws, row, COL_COMMENT)
        ng_reasons = []
        detail_ng_reasons = []  # codebeamer엔 안 보내고 요약 시트에만 보여줄 상세 사유

        # 저장 규칙
        save_result = check_save_rule(ws, row)
        if save_result is None:
            save_skip += 1
        else:
            is_ok, reasons, details = save_result
            if is_ok:
                ws.cell(row=row, column=COL_SAVE_RULE).value = 1
                save_ok += 1
            else:
                ws.cell(row=row, column=COL_SAVE_RULE).value = 2
                ng_reasons.extend(reasons)
                detail_ng_reasons.extend(details)
                save_ng += 1

        ver_ng_reasons = []
        ver_detail_reasons = []
        ver_checked = False

        # 버전 규칙
        ver_result = check_version_rule(ws, row)
        if ver_result is None:
            ver_skip += 1
        else:
            ver_checked = True
            is_ok, reasons, details = ver_result
            if is_ok:
                ver_ok += 1
            else:
                ver_ng_reasons.extend(reasons)
                ver_detail_reasons.extend(details)
                ver_ng += 1

        # 이벤트성 산출물 Create Date 검사
        event_result = check_eventbased_create_date(ws, row)
        if event_result is None:
            event_skip += 1
        else:
            ver_checked = True
            is_ok, reasons, details = event_result
            if is_ok:
                event_ok += 1
            else:
                ver_ng_reasons.extend(reasons)
                ver_detail_reasons.extend(details)
                event_ng += 1

        # 주기적 활동 산출물 Create Date 검사
        periodic_result = check_periodic_create_date(ws, row, periodic_cadence, periodic_anchor)
        if periodic_result is None:
            periodic_skip += 1
        else:
            ver_checked = True
            is_ok, reasons, details = periodic_result
            if is_ok:
                periodic_ok += 1
            else:
                ver_ng_reasons.extend(reasons)
                ver_detail_reasons.extend(details)
                periodic_ng += 1

        # T열 최종 기입: 검사된 항목이 있을 때만, NG 사유가 없으면 1 아니면 2
        if ver_checked:
            if ver_ng_reasons:
                ws.cell(row=row, column=COL_VERSION_RULE).value = 2
                ng_reasons.extend(ver_ng_reasons)
                detail_ng_reasons.extend(ver_detail_reasons)
            else:
                ws.cell(row=row, column=COL_VERSION_RULE).value = 1

        # 문서 이력 기술 규칙
        doc_hist_result = check_doc_history_rule(ws, row)
        if doc_hist_result is None:
            doc_hist_skip += 1
        else:
            is_ok, reasons, details = doc_hist_result
            if is_ok:
                ws.cell(row=row, column=COL_DOC_HISTORY_RULE).value = 1
                doc_hist_ok += 1
            else:
                ws.cell(row=row, column=COL_DOC_HISTORY_RULE).value = 2
                ng_reasons.extend(reasons)
                detail_ng_reasons.extend(details)
                doc_hist_ng += 1

        # 상태 규칙 (리뷰 대상 버전 규칙도 여기 합쳐서 같은 codebeamer 필드로 반영한다)
        status_ng_reasons = []
        status_detail_reasons = []
        status_checked = False

        status_result = check_status_rule(ws, row)
        if status_result is None:
            status_skip += 1
        else:
            status_checked = True
            is_ok, reasons, details = status_result
            if is_ok:
                status_ok += 1
            else:
                status_ng_reasons.extend(reasons)
                status_detail_reasons.extend(details)
                status_ng += 1

        review_ver_result = check_review_report_version_rule(ws, row)
        if review_ver_result is None:
            review_ver_skip += 1
            # Approved인데 리뷰 대상 버전을 자동으로 못 읽은 경우만 "판정 불가" 목록에 안내
            status_raw = get_cell_value(ws, row, COL_STATUS)
            fail_reason = get_cell_value(ws, row, COL_VERSION_CHECK_FAIL_REASON)
            if status_raw == 'Approved' and fail_reason:
                version_check_failures.append({
                    "트래커명": get_cell_value(ws, row, COL_TRACKER_NAME),
                    "사유": fail_reason,
                })
        else:
            status_checked = True
            is_ok, reasons, details = review_ver_result
            if is_ok:
                review_ver_ok += 1
            else:
                status_ng_reasons.extend(reasons)
                status_detail_reasons.extend(details)
                review_ver_ng += 1

        if status_checked:
            if status_ng_reasons:
                ws.cell(row=row, column=COL_STATUS_RULE).value = 2
                ng_reasons.extend(status_ng_reasons)
                detail_ng_reasons.extend(status_detail_reasons)
            else:
                ws.cell(row=row, column=COL_STATUS_RULE).value = 1

        # Comment 일괄 기입 (codebeamer로 나가는 간단한 버전)
        if ng_reasons:
            ws.cell(row=row, column=COL_COMMENT).value = append_comment(
                existing_comment, ng_reasons
            )

        # 상세 사유는 codebeamer에 보내지 않고, 요약 시트에서만 보여주기 위해 따로 보관
        if detail_ng_reasons:
            row_details[row] = detail_ng_reasons

    _build_summary_sheet(wb, ws, row_details)
    _build_version_check_failure_sheet(wb, version_check_failures)
    wb.save(file_path)

    print(f"\n완료: {file_path} 저장됨")
    print(f"  [저장 규칙 1.2)3)]             OK: {save_ok}건 / NG: {save_ng}건 / 스킵: {save_skip}건")
    print(f"  [버전 규칙 2.1)]               OK: {ver_ok}건 / NG: {ver_ng}건 / 스킵: {ver_skip}건")
    print(f"  [이벤트성 Create Date 2.2)]    OK: {event_ok}건 / NG: {event_ng}건 / 스킵: {event_skip}건")
    print(f"  [주기적 Create Date 2.3)]     OK: {periodic_ok}건 / NG: {periodic_ng}건 / 스킵: {periodic_skip}건")
    print(f"  [문서 이력 기술 규칙 3.2)]      OK: {doc_hist_ok}건 / NG: {doc_hist_ng}건 / 스킵: {doc_hist_skip}건")
    print(f"  [상태 규칙 4.1)2)]            OK: {status_ok}건 / NG: {status_ng}건 / 스킵: {status_skip}건")
    print(f"  [리뷰 대상 버전 규칙]          OK: {review_ver_ok}건 / NG: {review_ver_ng}건 / 스킵: {review_ver_skip}건 (판정 불가: {len(version_check_failures)}건)")


if __name__ == "__main__":
    file_path = sys.argv[1] if len(sys.argv) > 1 else str(DEFAULT_FILE)
    print(file_path)
    run(file_path)